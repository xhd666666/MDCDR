import csv
import math
import os
import pickle
import random
import timeit
from collections import defaultdict

import networkx as nx
import numpy as np
import pandas as pd
import scipy.sparse as sp
import torch
from matplotlib import pyplot as plt
from openpyxl.workbook import Workbook
from rdkit import Chem
from rdkit.Chem import AllChem, RDKFingerprint, MACCSkeys, rdMolDescriptors, rdmolops
from rdkit.Chem.Scaffolds import MurckoScaffold
from scipy.stats import pearsonr, spearmanr
from sklearn.cluster import MiniBatchKMeans
from sklearn.decomposition import PCA
from sklearn.metrics import mean_absolute_error, mean_squared_error, r2_score, roc_auc_score, \
	precision_score, recall_score, f1_score, accuracy_score, average_precision_score
from sklearn.model_selection import train_test_split, KFold
from torch.utils.data import Dataset, DataLoader
from torch_geometric.data import Data, Batch
from torch_geometric.nn import graclus, max_pool
import torch.nn.functional as F

from pubchemfp import GetPubChemFPs

def load_leave_mixed_split(df_gdsc, df_tcga, group_col='cell',
						   filter_cols=None,
						   tcga_test_group_ratio=0.8,
						   random_state=42):
	if filter_cols is None:
		filter_cols = [group_col]

	gdsc_unique_values = {}
	for col in filter_cols:
		if col in df_gdsc.columns and col in df_tcga.columns:
			gdsc_unique_values[col] = set(df_gdsc[col].unique())
		else:
			raise ValueError(f"Column '{col}' must exist in both GDSC and TCGA datasets simultaneously")

	mask = pd.Series(True, index=df_tcga.index)
	for col, unique_vals in gdsc_unique_values.items():
		mask = mask & (~df_tcga[col].isin(unique_vals))

	df_tcga_filtered = df_tcga[mask].reset_index(drop=True)

	unique_groups = df_tcga_filtered[group_col].unique()
	n_groups = len(unique_groups)

	test_groups, remaining_groups = train_test_split(
		unique_groups,
		test_size=tcga_test_group_ratio,
		random_state=random_state,
		shuffle=True
	)

	test_set = df_tcga_filtered[df_tcga_filtered[group_col].isin(test_groups)].reset_index(drop=True)

	remaining_tcga = df_tcga_filtered[df_tcga_filtered[group_col].isin(remaining_groups)].reset_index(drop=True)

	tcga_train_part, val_set = train_test_split(
		remaining_tcga,
		test_size=0.5,
		random_state=random_state,
		shuffle=True
	)

	tcga_train_part = tcga_train_part.reset_index(drop=True)
	val_set = val_set.reset_index(drop=True)

	train_set = pd.concat([df_gdsc.reset_index(drop=True), tcga_train_part], ignore_index=True)
	train_set = train_set.sample(frac=1, random_state=random_state).reset_index(drop=True)

	return train_set, val_set, test_set

def load_leave_hybrid(df, col_name, val_ratio=0.1, test_ratio=0.1, random_state=42):
	np.random.seed(random_state)
	entity_counts = df.groupby(col_name).size()
	entities = entity_counts.index.tolist()

	np.random.shuffle(entities)

	total_samples = len(df)
	target_test_samples = total_samples * test_ratio

	test_entities = []
	current_test_samples = 0

	for entity in entities:
		if current_test_samples >= target_test_samples:
			break
		test_entities.append(entity)
		current_test_samples += entity_counts[entity]

	train_val_entities = [e for e in entities if e not in test_entities]

	test_set = df[df[col_name].isin(test_entities)].reset_index(drop=True)
	train_val_df = df[df[col_name].isin(train_val_entities)].reset_index(drop=True)

	if len(train_val_df) > 0:
		adjusted_val_ratio = val_ratio / (1 - len(test_set) / total_samples)
		adjusted_val_ratio = min(adjusted_val_ratio, 0.99)

		train_set, val_set = train_test_split(
			train_val_df,
			test_size=adjusted_val_ratio,
			random_state=random_state
		)
		train_set = train_set.reset_index(drop=True)
		val_set = val_set.reset_index(drop=True)
	else:
		train_set = train_val_df
		val_set = pd.DataFrame(columns=df.columns)

	return train_set, val_set, test_set

def atom_features(atom):
	# 44 +11 +11 +11 +1
	return np.array(one_hot_encoding_unk(atom.GetSymbol(),
										 ['C', 'N', 'O', 'S', 'F', 'Si', 'P', 'Cl', 'Br', 'Mg', 'Na', 'Ca', 'Fe', 'As',
										  'Al', 'I', 'B', 'V', 'K', 'Tl', 'Yb', 'Sb', 'Sn', 'Ag', 'Pd', 'Co', 'Se',
										  'Ti', 'Zn', 'H', 'Li', 'Ge', 'Cu', 'Au', 'Ni', 'Cd', 'In', 'Mn', 'Zr', 'Cr',
										  'Pt', 'Hg', 'Pb', 'X']) +
					one_hot_encoding(atom.GetDegree(), [0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10]) +
					one_hot_encoding(atom.GetTotalNumHs(), [0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10]) +
					one_hot_encoding(atom.GetImplicitValence(), [0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10]) +
					[atom.GetIsAromatic()])

def one_hot_encoding(x, allowable_set):
	if x not in allowable_set:
		raise Exception('input {0} not in allowable set{1}:'.format(x, allowable_set))
	return list(map(lambda s: x == s, allowable_set))

# one ont encoding with unknown symbol
def one_hot_encoding_unk(x, allowable_set):
	if x not in allowable_set:
		x = allowable_set[-1]
	return list(map(lambda s: x == s, allowable_set))

def smile_to_graph(smile):
	mol = Chem.MolFromSmiles(smile)
	mol_size = mol.GetNumAtoms()

	mol_features = []
	for atom in mol.GetAtoms():
		feature = atom_features(atom)
		mol_features.append(feature / sum(feature))
	edges = []

	for bond in mol.GetBonds():
		edges.append([bond.GetBeginAtomIdx(), bond.GetEndAtomIdx()])
	g = nx.Graph(edges).to_directed()
	edge_index = []
	for e1, e2 in g.edges:
		edge_index.append([e1, e2])
	return mol_size, mol_features, edge_index

def smileToECFP4(smile):
	mol = Chem.MolFromSmiles(smile)
	morgan_fp = AllChem.GetMorganFingerprintAsBitVect(mol, radius=2, nBits=2048, useChirality=False)
	return morgan_fp.ToBitString()

def smileToESPF(smile):
	mol = Chem.MolFromSmiles(smile)
	morgan_fp = AllChem.GetMorganFingerprintAsBitVect(mol, radius=2, nBits=2048, useChirality=True)
	return morgan_fp.ToBitString()

def smileToPubchem(smile):
	mol = Chem.MolFromSmiles(smile)
	pubchem_fp = GetPubChemFPs(mol)
	return pubchem_fp

def load_gene_expression(filepath):
	cell_map = {}
	expression_data = pd.read_csv(filepath, sep=',', header=0, index_col=[0])
	expression_array = expression_data.to_numpy()
	idx = 0
	for cell in expression_data.index:
		cell_map[cell] = expression_array[idx]
		idx = idx + 1
	return cell_map

def load_pathway(filepath):
	pathway_map = {}
	pathway_data = pd.read_csv(filepath, sep=',', header=0, index_col=[0])
	pathway_data = pathway_data.T
	pathway_array = pathway_data.to_numpy()
	idx = 0
	for cell in pathway_data.index:
		pathway_map[cell] = pathway_array[idx]
		idx = idx + 1
	return pathway_map

def normalize_rows_minmax(df):

	row_mins = df.min(axis=1)
	row_maxs = df.max(axis=1)
	row_ranges = row_maxs - row_mins

	normalized_df = (df.sub(row_mins, axis=0)
					 .div(row_ranges, axis=0))

	return normalized_df

def eval_predict(y_label, y_pred):
	mae = mean_absolute_error(y_label, y_pred)
	mse = mean_squared_error(y_label, y_pred)
	rmse = np.sqrt(mse)
	r2 = r2_score(y_label, y_pred)
	pearson = pearsonr(y_label, y_pred)[0]
	pearson_p_value = pearsonr(y_label, y_pred)[1]
	spearman = spearmanr(y_label, y_pred)[0]
	spearman_p_value = spearmanr(y_label, y_pred)[1]
	return mse, rmse, mae, r2, pearson, pearson_p_value, spearman, spearman_p_value

def metrics_cls(y_true, y_predict):
	predicted_labels = list(map(lambda x: np.argmax(x), y_predict))
	predicted_scores = list(map(lambda x: x[1], y_predict))
	t, s, y = [], [], []
	for j in range(len(y_true)):
		t.append(y_true[j])
		y.append(predicted_labels[j])
		s.append(predicted_scores[j])
	auc = roc_auc_score(t, s)
	aupr = average_precision_score(t, s)
	precision = precision_score(t, y)
	recall = recall_score(t, y)
	f1 = f1_score(t, y)
	acc = accuracy_score(t, y)
	return auc, aupr, precision, recall, f1, acc

class MyDataset(Dataset):
	def __init__(self, drug_dict, cell_dict, res):
		super(MyDataset, self).__init__()
		self.drug, self.cell = drug_dict, cell_dict
		res.reset_index(drop=True, inplace=True)
		self.drug_name = res['drug']
		self.Cell_line_name = res['cell']
		self.value = res['res']

	def __len__(self):
		return len(self.value)

	def __getitem__(self, index):
		pubchem = self.drug_name[index]
		depmap = self.Cell_line_name[index]
		return self.drug[pubchem], self.cell[depmap], self.value[index], pubchem, depmap


def load_drug_dict(drug_path):
	drug_dict = {}
	drugs_pubchem_smiles = pd.read_csv(drug_path, sep=',')
	for idx in drugs_pubchem_smiles.index:
		pubchem = int(drugs_pubchem_smiles.loc[idx, 'pubchem_id'])
		smiles = drugs_pubchem_smiles.loc[idx, 'isomeric_smiles']
		c_size, features, edge_index = smile_to_graph(smiles)
		morgan = smileToECFP4(smiles)
		morgan = [int(c) for c in morgan]
		espf = smileToESPF(smiles)
		espf = [int(c) for c in espf]
		pubchem_fp = smileToPubchem(smiles)
		gin = smile_to_gin(pubchem)
		graph = Data(x=torch.tensor(features, dtype=torch.float32),
					 edge_index=torch.tensor(edge_index, dtype=torch.int64).T,
					 morgan=torch.tensor(morgan, dtype=torch.float32),
					 espf=torch.tensor(espf, dtype=torch.float32),
					 pubchem_fp=torch.tensor(pubchem_fp, dtype=torch.float32),
					 pubchem=pubchem,
					 gin=gin,
					 smiles=smiles)
		drug_dict[pubchem] = graph
	return drug_dict

def smile_to_gin(smile):
	save_dir = "data/gin"
	feature_path = os.path.join(save_dir, f"{smile}.pkl")
	with open(feature_path, "rb") as f:
		feature = pickle.load(f)
	mean_feat = torch.mean(feature, dim=0)
	max_feat = torch.max(feature, dim=0)[0]
	min_feat = torch.min(feature, dim=0)[0]
	# feature = torch.cat([mean_feat, max_feat, min_feat])
	return mean_feat

def load_cell_dict(exp_path, pathway_path):
	cell_map = load_gene_expression(exp_path)
	pathway_map = load_pathway(pathway_path)
	edge_index = load_edge_index(exp_path)
	cell_dict = {}
	for cell in cell_map.keys():
		gene_expression = cell_map[cell]
		pathway = pathway_map[cell]
		x = torch.tensor(gene_expression, dtype=torch.float32)
		mean = x.mean()
		std = x.std()
		x = (x - mean) / std
		x = x.unsqueeze(1)
		graph = Data(x=x, edge_index = torch.tensor(edge_index, dtype=torch.long),
					 gene_expression = torch.tensor(gene_expression, dtype=torch.float32),
					 pathway = torch.tensor(pathway, dtype=torch.float32))
		cell_dict[cell] = graph
	return cell_dict

def _collate(samples):
	drugs, cells, res, pubchem, depmap = map(list, zip(*samples))

	morgan = [drug.morgan for drug in drugs]
	morgan = torch.stack(morgan)
	espf = [drug.espf for drug in drugs]
	espf = torch.stack(espf)
	pubchem_fp = [drug.pubchem_fp for drug in drugs]
	pubchem_fp = torch.stack(pubchem_fp)
	gin = [drug.gin for drug in drugs]
	gin = torch.stack(gin)

	gene_expression = [cell.gene_expression for cell in cells]
	gene_expression = torch.stack(gene_expression)
	pathway = [cell.pathway for cell in cells]
	pathway = torch.stack(pathway)

	batched_drug = Batch.from_data_list(drugs)
	batched_cell = Batch.from_data_list(cells)

	batched_drug.morgan = morgan
	batched_drug.espf = espf
	batched_drug.pubchem_fp = pubchem_fp
	batched_drug.gin = gin

	batched_cell.gene_expression = gene_expression
	batched_cell.pathway = pathway

	return batched_drug, batched_cell, torch.tensor(res, dtype=torch.int64), pubchem, depmap

def _collate_regr(samples):
	drugs, cells, res, pubchem, depmap = map(list, zip(*samples))

	morgan = [drug.morgan for drug in drugs]
	morgan = torch.stack(morgan)
	espf = [drug.espf for drug in drugs]
	espf = torch.stack(espf)
	pubchem_fp = [drug.pubchem_fp for drug in drugs]
	pubchem_fp = torch.stack(pubchem_fp)
	gin = [drug.gin for drug in drugs]
	gin = torch.stack(gin)

	gene_expression = [cell.gene_expression for cell in cells]
	gene_expression = torch.stack(gene_expression)
	pathway = [cell.pathway for cell in cells]
	pathway = torch.stack(pathway)

	batched_drug = Batch.from_data_list(drugs)
	batched_cell = Batch.from_data_list(cells)

	batched_drug.morgan = morgan
	batched_drug.espf = espf
	batched_drug.pubchem_fp = pubchem_fp
	batched_drug.gin = gin

	batched_cell.gene_expression = gene_expression
	batched_cell.pathway = pathway

	return batched_drug, batched_cell, torch.tensor(res, dtype=torch.float32), pubchem, depmap


def train(args, model, train_loader, val_loader, test_loader, fold, mode):
	best_auc = 0
	stop_count = 0

	wb = Workbook()
	ws1 = wb.active
	ws1.title = "fusion"
	ws2 = wb.create_sheet("context")
	ws3 = wb.create_sheet("structure")

	print('start')
	for epoch in range(args.epochs):
		start = timeit.default_timer()
		model.train(train_loader)
		row1, row2, row3, auc = model.validate(val_loader)
		end = timeit.default_timer()
		time = end - start
		ws1.append(row1)
		ws2.append(row2)
		ws3.append(row3)
		print(epoch, time)
		print('row1=', row1)
		print('row2=', row2)
		print('row3=', row3)
		if best_auc < auc:
			best_auc = auc
			stop_count = 0
		else:
			stop_count = stop_count + 1

		if stop_count == args.patience:
			break
		print('none update count: ', stop_count)
	print(f"\nFold {fold} 训练完成，加载最佳模型评估测试集...")
	model.load_state_dict(torch.load(mode + '_model_' + str(fold) + '.pth'))
	row1, row2, row3, _ = model.validate(test_loader)
	print('row1=', row1)
	print('row2=', row2)
	print('row3=', row3)
	ws1.append([])
	ws1.append(['Val AUC', 'Val AUPR', 'Val Precision', 'Val Recall', 'Val F1', 'Val Acc'])
	ws1.append(row1)
	ws2.append([])
	ws2.append(['Val AUC', 'Val AUPR', 'Val Precision', 'Val Recall', 'Val F1', 'Val Acc'])
	ws2.append(row2)
	ws3.append([])
	ws3.append(['Val AUC', 'Val AUPR', 'Val Precision', 'Val Recall', 'Val F1', 'Val Acc'])
	ws3.append(row3)
	wb.save(mode + ':out' + str(fold) + '.xlsx')

def train_regr(args, model, train_loader, val_loader, test_loader, fold, mode):
	best_val_mse = 100

	stop_count = 0

	wb = Workbook()
	ws1 = wb.active
	ws1.title = "fusion"
	ws1.append(['Val MSE', 'Val RMSE', 'Val MAE', 'Val R2', 'Val Pearson', 'Val Spearman'])
	ws2 = wb.create_sheet("context")
	ws3 = wb.create_sheet("structure")

	print('start')
	for epoch in range(args.epochs):
		start = timeit.default_timer()
		model.train(train_loader)
		row1, row2, row3, mse = model.validate(val_loader)
		end = timeit.default_timer()
		time = end - start
		ws1.append(row1)
		ws2.append(row2)
		ws3.append(row3)
		print(epoch, time)
		print('row1=', row1)
		print('row2=', row2)
		print('row3=', row3)
		if best_val_mse > mse:
			best_val_mse = mse
			stop_count = 0
			model.save_model(mode, fold)
		else:
			stop_count = stop_count + 1

		print('none update count: ', stop_count)

	print(f"\nFold {fold} training completed, loading the best model to evaluate the test set...")
	model.load_state_dict(torch.load(mode + '_model_' + str(fold) + '.pth'))
	row1, row2, row3, mse = model.validate(test_loader)
	print('row1=', row1)
	print('row2=', row2)
	print('row3=', row3)
	ws1.append([])
	ws1.append(['Test MSE', 'Test RMSE', 'Test MAE', 'Test R2', 'Test Pearson', 'Test Spearman'])
	ws1.append(row1)
	ws2.append([])
	ws2.append(['Test MSE', 'Test RMSE', 'Test MAE', 'Test R2', 'Test Pearson', 'Test Spearman'])
	ws2.append(row2)
	ws3.append([])
	ws3.append(['Test MSE', 'Test RMSE', 'Test MAE', 'Test R2', 'Test Pearson', 'Test Spearman'])
	ws3.append(row3)
	wb.save(mode+':out'+str(fold)+'.xlsx')

def load_edge_index(exp_path):
	exp = pd.read_csv(exp_path, index_col=0)
	gene_list = exp.columns.to_list()
	ensp_map = ensp_to_hugo_map()
	edges = pd.read_csv('data/STRING/9606.protein.links.v11.5.txt', sep=' ')
	selected_edges = edges['combined_score'] > 800
	edge_list = edges[selected_edges][["protein1", "protein2"]].values.tolist()
	edge_list = [[ensp_map[edge[0]], ensp_map[edge[1]]] for edge in edge_list if
			 edge[0] in ensp_map.keys() and edge[1] in ensp_map.keys()]
	edge_index = []
	for i in edge_list:
		if (i[0] in gene_list) & (i[1] in gene_list):
			edge_index.append((gene_list.index(i[0]), gene_list.index(i[1])))
			edge_index.append((gene_list.index(i[1]), gene_list.index(i[0])))
	edge_index = list(set(edge_index))
	edge_index = np.array(edge_index, dtype=np.int64).T
	return edge_index

def ensp_to_hugo_map():
	with open('data/STRING/9606.protein.info.v11.5.txt') as csv_file:
		next(csv_file)  # Skip first line
		csv_reader = csv.reader(csv_file, delimiter='\t')
		ensp_map = {row[0]: row[1] for row in csv_reader if row[0] != ""}

	return ensp_map

def get_predefine_cluster(exp_path, device, save_path="cluster_predefine.pth"):
	if os.path.exists(save_path):
		print(f"Local cache file detected: {save_path}, loading...")
		cluster_predefine = torch.load(save_path, map_location=device)
		print("Cluster results loaded successfully!")
		return cluster_predefine
	edge_index = load_edge_index(exp_path)
	exp = pd.read_csv(exp_path, index_col=0)
	g = Data(edge_index=torch.tensor(edge_index, dtype=torch.long), x=torch.zeros(len(exp.columns.to_list()), 1))
	g = Batch.from_data_list([g])
	cluster_predefine = {}
	for i in range(5):
		cluster = graclus(g.edge_index, None, g.x.size(0))
		# print(len(cluster.unique()))
		g = max_pool(cluster, g, transform=None)
		cluster_predefine[i] = cluster
	cluster_predefine = {i: j.to(device) for i, j in cluster_predefine.items()}

	save_dir = os.path.dirname(save_path)
	if save_dir and not os.path.exists(save_dir):
		os.makedirs(save_dir, exist_ok=True)

	torch.save(cluster_predefine, save_path)
	print(f"Cluster results saved to: {save_path}")
	return cluster_predefine

def orthogonality_loss(s_z, p_z):
	s_l2_norm = torch.norm(s_z, p=2, dim=1, keepdim=True).detach()
	s_l2 = s_z.div(s_l2_norm.expand_as(s_z) + 1e-6)

	p_l2_norm = torch.norm(p_z, p=2, dim=1, keepdim=True).detach()
	p_l2 = p_z.div(p_l2_norm.expand_as(p_z) + 1e-6)

	ortho_loss = torch.mean((s_l2.t().mm(p_l2)).pow(2))
	return ortho_loss

def generate_mask(x, mask_rate = 0.15):
	device = x.device
	B, H = x.shape
	mask = torch.ones(B, H, device=device)
	num_mask = int(H * mask_rate)
	mask_indices = torch.rand(B, H, device=device).topk(num_mask, dim=1).indices
	mask.scatter_(1, mask_indices, 0)
	return mask.unsqueeze(-1) if x.dim() == 3 else mask

def calc_masked_loss(pred, target, mask):
	num_masked = (1 - mask).sum()
	if num_masked == 0:
		return torch.tensor(0.0, device=pred.device)

	loss = F.mse_loss(pred * (1 - mask), target * (1 - mask), reduction='sum')
	return loss / num_masked
